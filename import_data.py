from pathlib import Path
from openpyxl import load_workbook
from database import get_connection


BASE_DIR = Path(__file__).resolve().parent
RESOURCES_DIR = BASE_DIR / "resources"


def _resolve_resource_path(*candidates):
    for name in candidates:
        candidate = RESOURCES_DIR / name
        if candidate.exists():
            return candidate

    available_files = sorted([p.name for p in RESOURCES_DIR.glob("*.xlsx")])
    raise FileNotFoundError(
        "Required Excel file not found. Tried: "
        f"{', '.join(candidates)}. "
        f"Available .xlsx files in resources: {available_files}"
    )


def import_if_empty():
    conn = get_connection()
    cursor = conn.cursor()
    count = cursor.execute("SELECT COUNT(*) FROM product").fetchone()[0]
    conn.close()

    if count == 0:
        print("Database is empty. Importing data from Excel files...")
        try:
            import_all()
            print("Data import completed successfully!")
        except FileNotFoundError as error:
            print("Data import skipped: missing Excel file.")
            print(error)
            print("Add required files to the 'resources' folder")
    else:
        print(f"Database already contains {count} products. Skipping import.")


def import_all():
    import_pickup_points()
    import_users()
    import_products()
    import_orders()


def import_pickup_points():
    print("Importing pickup points...")
    pickup_file = _resolve_resource_path(
        "Пункты выдачи_import.xlsx"
    )
    wb = load_workbook(pickup_file)
    ws = wb.active

    conn = get_connection()
    cursor = conn.cursor()

    count = 0
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0]:
            address = str(row[0]).strip()
            cursor.execute(
                "INSERT OR IGNORE INTO pickup_point (address) VALUES (?)",
                (address,)
            )
            count += 1

    conn.commit()
    conn.close()
    wb.close()
    print(f"Imported {count} pickup points")


def import_users():
    print("Importing users...")
    user_file = _resolve_resource_path("user_import.xlsx")
    wb = load_workbook(user_file)
    ws = wb.active

    conn = get_connection()
    cursor = conn.cursor()

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            role_name = str(row[0]).strip()
            full_name = str(row[1]).strip()
            login = str(row[2]).strip()
            password = str(row[3]).strip()

            cursor.execute(
                "INSERT OR IGNORE INTO role (role_name) VALUES (?)",
                (role_name,)
            )
            cursor.execute(
                "SELECT role_id FROM role WHERE role_name = ?",
                (role_name,)
            )
            role_id = cursor.fetchone()[0]

            cursor.execute("""
                INSERT OR IGNORE INTO
                           user (role_id, full_name, login, password)
                VALUES (?, ?, ?, ?)
            """, (role_id, full_name, login, password))
            count += 1

    conn.commit()
    conn.close()
    wb.close()
    print(f"Imported {count} users")


def import_products():
    print("Importing products...")
    product_file = _resolve_resource_path("Tovar.xlsx")
    wb = load_workbook(product_file)
    ws = wb.active

    conn = get_connection()
    cursor = conn.cursor()

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            article = str(row[0]).strip()
            name = str(row[1]).strip()
            unit = str(row[2]).strip() if row[2] else 'шт.'
            price = float(row[3]) if row[3] else 0.0
            supplier_name = str(row[4]).strip()
            manufacturer_name = str(row[5]).strip()
            category_name = str(row[6]).strip()
            discount = int(row[7]) if row[7] else 0
            stock_qty = int(row[8]) if row[8] else 0
            description = str(row[9]).strip() if row[9] else None
            photo_filename = str(row[10]).strip() if row[10] else None

            cursor.execute(
                "INSERT OR IGNORE INTO supplier (supplier_name) VALUES (?)",
                (supplier_name,)
            )
            cursor.execute(
                "SELECT supplier_id FROM supplier WHERE supplier_name = ?",
                (supplier_name,)
            )
            supplier_id = cursor.fetchone()[0]

            cursor.execute(
                """INSERT OR IGNORE INTO manufacturer (manufacturer_name)
                    VALUES (?)""",
                (manufacturer_name,)
            )
            cursor.execute(
                """SELECT manufacturer_id FROM manufacturer
                    WHERE manufacturer_name = ?""",
                (manufacturer_name,)
            )
            manufacturer_id = cursor.fetchone()[0]

            cursor.execute(
                "INSERT OR IGNORE INTO category (category_name) VALUES (?)",
                (category_name,)
            )
            cursor.execute(
                "SELECT category_id FROM category WHERE category_name = ?",
                (category_name,)
            )
            category_id = cursor.fetchone()[0]

            photo_path = None
            if photo_filename:
                photo_path = f"static/images/{photo_filename}"

            cursor.execute("""
                INSERT OR IGNORE INTO product (
                    article, name, unit, price, supplier_id, manufacturer_id,
                    category_id, discount, stock_qty, description, photo_path
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                article, name, unit, price, supplier_id, manufacturer_id,
                category_id, discount, stock_qty, description, photo_path
            ))
            count += 1

    conn.commit()
    conn.close()
    wb.close()
    print(f"Imported {count} products")


def import_orders():
    print("Importing orders...")
    order_file = _resolve_resource_path("Заказ_import.xlsx")
    wb = load_workbook(order_file)
    ws = wb.active

    conn = get_connection()
    cursor = conn.cursor()

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            order_id = int(row[0])
            articles = str(row[1]).strip()
            order_date = str(row[2]).strip()
            delivery_date = str(row[3]).strip()
            pickup_point_id = int(row[4])
            full_name = str(row[5]).strip()
            pickup_code = str(row[6]).strip()
            status = str(row[7]).strip() if row[7] else 'Новый'

            cursor.execute(
                "SELECT user_id FROM user WHERE full_name = ?",
                (full_name,)
            )
            user_result = cursor.fetchone()

            if user_result:
                user_id = user_result[0]

                cursor.execute("""
                    INSERT OR IGNORE INTO "order" (
                        order_id, articles, order_date, delivery_date,
                        pickup_point_id, user_id, pickup_code, status
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    order_id, articles, order_date, delivery_date,
                    pickup_point_id, user_id, pickup_code, status
                ))
                count += 1
            else:
                print(f"User '{full_name}' not found for order {order_id}")

    conn.commit()
    conn.close()
    wb.close()
    print(f"Imported {count} orders")
